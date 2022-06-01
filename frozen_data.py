# -*- coding: utf-8 -*-
"""
Created on Mon Mar  7 16:55:08 2022

@author: VK
"""

# функция по обработке новых данных заморозки
# принимает название файла, путь к файлу, название листа
def fill_frozen_sheet(wbName, wbPath, wsData):
    from openpyxl import load_workbook


# секция открытия файл для работы
    wb = load_workbook(wbPath+wbName)
    ws = wb[wsData]

    i = 2                   # данные находятся начиная со второго ряда
    space = 72              # расстояние между первыми рядами исходных и обработанных данных
    SourceRow = i
    InRow = SourceRow + space
    # добавление шапки и вспомогательных значений в первом блоке
    ws.cell(column=3, row=InRow-4).value = "23:59:00"
    ws.cell(column=3, row=InRow-3).value = "0:01:00"
    ws.cell(column=3, row=InRow-2).value = "0:02:00"
    ws.cell(column=2, row=InRow-1).value = "чистое"
    ws.cell(column=3, row=InRow-1).value = "грязное"
    ws.cell(column=4, row=InRow-1).value = "чистое"
    ws.cell(column=5, row=InRow-1).value = "грязное"
    ws.cell(column=6, row=InRow-1).value = "в минутах"
    
    # добавление первой строки - там другие формулы
    ws.cell(column=1, row=InRow).value = "=A" + str(SourceRow)
    ws.cell(column=2, row=InRow).number_format = "h:mm"
    ws.cell(column=2, row=InRow).value = "=C" + str(SourceRow) + "-B" + str(SourceRow)
    ws.cell(column=3, row=InRow).number_format = "h:mm"
    ws.cell(column=3, row=InRow).value = "=C" + str(SourceRow) + "-B" + str(SourceRow)
    ws.cell(column=4, row=InRow).number_format = "h:mm"
    ws.cell(column=4, row=InRow).value = "=E" + str(SourceRow) + "-D" + str(SourceRow)
    ws.cell(column=5, row=InRow).number_format = "h:mm"
    ws.cell(column=5, row=InRow).value = "=E" + str(SourceRow) + "-D" + str(SourceRow)
    ws.cell(column=6, row=InRow).value = "=HOUR(D" + str(InRow) + ")*60+MINUTE(D" + str(InRow) + ")"
    ws.cell(column=7, row=InRow).value = "=HOUR(E" + str(InRow) + ")*60+MINUTE(E" + str(InRow) + ")"
    
    # добавление основного массива формул
    i = 3
    SourceRow = i
    while (ws.cell(column=1, row=SourceRow).value != None) : 
        InRow = SourceRow + space
        # обычные формулы
        ws.cell(column=1, row=InRow).value = "=A" + str(SourceRow)
        ws.cell(column=2, row=InRow).number_format = "h:mm"
        ws.cell(column=2, row=InRow).value = "=C" + str(SourceRow) + "-B" + str(SourceRow)
        ws.cell(column=3, row=InRow).number_format = "h:mm"
        ws.cell(column=3, row=InRow).value = "=C" + str(SourceRow) + "-C" + str(SourceRow-1)
        ws.cell(column=4, row=InRow).number_format = "h:mm"
        ws.cell(column=4, row=InRow).value = "=E" + str(SourceRow) + "-D" + str(SourceRow)
        ws.cell(column=5, row=InRow).number_format = "h:mm"
        ws.cell(column=5, row=InRow).value = "=E" + str(SourceRow) + "-E" + str(SourceRow-1)
        ws.cell(column=6, row=InRow).value = "=HOUR(D" + str(InRow) + ")*60+MINUTE(D" + str(InRow) + ")"
        ws.cell(column=7, row=InRow).value = "=HOUR(E" + str(InRow) + ")*60+MINUTE(E" + str(InRow) + ")"
        
        # корректировка формул при переходе через полночь
        if (ws.cell(column=3, row=SourceRow).value.hour < ws.cell(column=2, row=SourceRow).value.hour) :
            ws.cell(column=2, row=InRow).value += "+$C$70-$C$71+$C$72"
        if (ws.cell(column=3, row=SourceRow).value.hour < ws.cell(column=3, row=SourceRow-1).value.hour) :
            ws.cell(column=3, row=InRow).value += "+$C$70-$C$71+$C$72"
        if (ws.cell(column=5, row=SourceRow).value.hour < ws.cell(column=4, row=SourceRow).value.hour) :
            ws.cell(column=4, row=InRow).value += "+$C$70-$C$71+$C$72"
        if (ws.cell(column=5, row=SourceRow).value.hour < ws.cell(column=5, row=SourceRow-1).value.hour) :
            ws.cell(column=5, row=InRow).value += "+$C$70-$C$71+$C$72"
        i += 1
        SourceRow = i
    LastSoRow = i-1
    LastInRow = LastSoRow+space

    # добавление нижнего блока
    ws["B140"] = "вакуумация"
    ws["B141"] = "чистое"
    ws["C141"] = "грязное"
    ws["D140"] = "заморозка"
    ws["D141"] = "чистое"
    ws["E141"] = "грязное"
    ws["A142"] = "среднее"
    ws["A143"] = "мин"
    ws["A144"] = "макс"
    ws["B142"].number_format = "h:mm"
    ws["B142"] = "=AVERAGE(B74:B" + str(LastInRow) + ")"
    ws["C142"].number_format = "h:mm"
    ws["C142"] = "=AVERAGE(C74:C" + str(LastInRow) + ")"
    ws["D142"].number_format = "h:mm"
    ws["D142"] = "=AVERAGE(D74:D" + str(LastInRow) + ")"
    ws["E142"].number_format = "h:mm"
    ws["E142"] = "=AVERAGE(E74:E" + str(LastInRow) + ")"
    ws["B143"].number_format = "h:mm"
    ws["B143"] = "=MIN(B74:B" + str(LastInRow) + ")"
    ws["C143"].number_format = "h:mm"
    ws["C143"] = "=MIN(C74:C" + str(LastInRow) + ")"
    ws["D143"].number_format = "h:mm"
    ws["D143"] = "=MIN(D74:D" + str(LastInRow) + ")"
    ws["E143"].number_format = "h:mm"
    ws["E143"] = "=MIN(E74:E" + str(LastInRow) + ")"
    ws["B144"].number_format = "h:mm"
    ws["B144"] = "=MAX(B74:B" + str(LastInRow) + ")"
    ws["C144"].number_format = "h:mm"
    ws["C144"] = "=MAX(C74:C" + str(LastInRow) + ")"
    ws["D144"].number_format = "h:mm"
    ws["D144"] = "=MAX(D74:D" + str(LastInRow) + ")"
    ws["E144"].number_format = "h:mm"
    ws["E144"] = "=MAX(E74:E" + str(LastInRow) + ")"
    ws["F142"] = "<21"
    ws["G142"] = "<21"
    ws["F143"] = "=COUNTIF(F74:F" + str(LastInRow) + ",F142)"
    ws["G143"] = "=COUNTIF(G74:G" + str(LastInRow) + ",G142)"
    ws["F144"] = "<26"
    ws["G144"] = "<26"
    ws["F145"] = "=COUNTIF(F74:F" + str(LastInRow) + ",F144)"
    ws["G145"] = "=COUNTIF(G74:G" + str(LastInRow) + ",G144)"
    ws["F146"] = "<31"
    ws["G146"] = "<31"
    ws["F147"] = "=COUNTIF(F74:F" + str(LastInRow) + ",F146)"
    ws["G147"] = "=COUNTIF(G74:G" + str(LastInRow) + ",G146)"
    
    ws["E150"] = "всего партий"
    ws["E151"] = "средний размер партии, кг"
    ws["E152"] = "заморозка 20 мин. и меньше"
    ws["E154"] = "заморозка 25 мин. и меньше"
    ws["E156"] = "заморозка 30 мин. и меньше"
    ws["E159"] = "средняя производительность, кг/час"
    ws["E160"] = "производительность по чистому"
    ws["G150"] = "=MAX(A2:A" + str(LastSoRow) + ")"
    ws["G151"] = "=AVERAGE(J2:J" + str(LastSoRow) + ")"
    ws["G152"] = "=F143"
    ws["G153"] = "=G152/G150"
    ws["G153"].number_format = "0.0%"
    ws["G154"] = "=F145"
    ws["G155"] = "=G154/G150"
    ws["G155"].number_format = "0.0%"
    ws["G156"] = "=F147"
    ws["G157"] = "=G156/G150"
    ws["G157"].number_format = "0.0%"
    ws["G159"] = "=MAX(K2:K" + str(LastSoRow) + ")/SUM(G74:G" + str(LastInRow) + ")*60"
    ws["G160"] = "=MAX(K2:K" + str(LastSoRow) + ")/SUM(F74:F" + str(LastInRow) + ")*60"

# сохранить изменения, внесенные в файл
    wb.save(wbPath+wbName)
    return None

def add_charts (wbName, wbPath, wsData):
    """ добавление стандартных диаграмм на лист"""
    from openpyxl import load_workbook
    from openpyxl.chart import (
    LineChart,
    Reference,
    Series
)
    from openpyxl.drawing.text import (
        Paragraph, 
        ParagraphProperties, 
        CharacterProperties, 
        RichTextProperties, 
        Font,
        RegularTextRun
)

    # секция открытия файл для работы
    wb = load_workbook(wbPath+wbName)
    ws = wb[wsData]

    # определение первого и последнего ряда с данными для графиков
    StartRow = EndRow = 74
    while (ws["A"+str(EndRow)].value != None) : 
        EndRow += 1
    EndRow -= 1
    
# добавление диаграммы по времени вакуумации
    c1 = LineChart()
    c1.height = 10
    c1.width = 17
    c1.title = "Время вакуумации, чч:мин"
    c1.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(latin=Font(typeface='Calibri'), sz=1400, b=False))
    c1.style = 13
    c1.legend.position = "b"
    ws.add_chart(c1, "I139")

    # добавление рядов с данными в диаграмму
    values = Reference(ws, min_col=2, min_row=StartRow, max_row=EndRow)
    series = Series(values, title="чистое")
    c1.append(series)
    values = Reference(ws, min_col=3, min_row=StartRow, max_row=EndRow)
    series = Series(values, title="грязное")
    c1.append(series)
    s11 = c1.series[0]
    s11.graphicalProperties.line.solidFill = "0099CC"
    s12 = c1.series[1]
    s12.graphicalProperties.line.solidFill = "EE9900"
# ----------- конец первой диаграммы    
    
# добавление диаграммы по времени заморозки
    c2 = LineChart()
    c2.height = 10
    c2.width = 17
    c2.title = "Время заморозки, чч:мин"
    c2.title.text.rich.paragraphs[0].pPr = ParagraphProperties(defRPr=CharacterProperties(latin=Font(typeface='Calibri'), sz=1400, b=False))
    c2.style = 13
    c2.legend.position = "b"
    ws.add_chart(c2, "I159")

    # добавление рядов с данными в диаграмму
    values = Reference(ws, min_col=4, min_row=StartRow, max_row=EndRow)
    series = Series(values, title="чистое")
    c2.append(series)
    values = Reference(ws, min_col=5, min_row=StartRow, max_row=EndRow)
    series = Series(values, title="грязное")
    c2.append(series)
    s21 = c2.series[0]
    s21.graphicalProperties.line.solidFill = "0099CC"
    s22 = c2.series[1]
    s22.graphicalProperties.line.solidFill = "EE9900"
# ----------- конец второй диаграммы

# сохранить изменения, внесенные в файл
    wb.save(wbPath+wbName)
    return None

def test(wbName, wbPath, wsData):
    from openpyxl import load_workbook


# секция открытия файл для работы
    wb = load_workbook(wbPath+wbName)
    ws = wb[wsData]
#    ws.cell(column=11, row=4).value = "5"

# сохранить изменения, внесенные в файл
    wb.save(wbPath+wbName)
    return None


# -------- ИСПОЛНЯЕМЫЙ КОД -------------
# данные файла и листа для обработки
# !!! не забыть перевернуть слэши в пути!!!!
ExFile = "расчеты заморозка обработанные.xlsx"
ExPath = "I:/ФБП/Заморозка/Модель заморозки/"
ExSheet = "24-25.03"

# уточнение что именно будет исполняться
check = input(f"Сейчас будет выполнена обработка следующего листа!!!\nФайл: {ExFile}\nПуть: {ExPath}\nЛист: {ExSheet}\nПродолжить? y/n ")
if check=="y" :
    fill_frozen_sheet(ExFile, ExPath, ExSheet)
    add_charts(ExFile, ExPath, ExSheet )
    print ("Обработано успешно")
else :
    print ("Отказ от обработки")










